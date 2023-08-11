"""
Example script for testing the Forest theme

Author: rdbende
License: MIT license
Source: https://github.com/rdbende/ttk-widget-factory
"""

import tkinter as tk
from tkinter import ttk
from tkinter import *
from openpyxl import *
import openpyxl


# IDEAS A PROBAR
# PONER 2 LISTBOX, UNA PARA EL NEGOCIO DE DONDE PROVIENE EL PRODUCTO A COMPRAR
# Y LA OTRA LA BASE DE DATOS DE DONDE LA EXTRAE
#
# AÑADIR DEL WIDGET # BASE DE DATOS # AL TREEVIEW
#
# AÑADIR EL WHATSAPP
#


# JERARQUIA
#
# ORDEN #XXX
#   TIENDA A        
#       PRODUCTO A  $XXX
#       PRODUCTO B  $XXX
#       PRODUCTO C  $XXX
#       PRODUCTO D  $XXX
#-------------------------SUBTOTAL
#                   $XXXA
##  TIENDA B     
#       PRODUCTO A  $XXX
#       PRODUCTO B  $XXX
#-------------------------SUBTOTAL
#                   $XXXB
#-------------------------TOTAL
#                   $XXXT
#
#   



# Update the listbox
def update(data):
    # Clear the listbox
    # limpia la listbox
    my_list.delete(0,END)

    # Add lreyes to listbox
    # se añaden los elementos de una lista en la listbox
    for item in data:
        my_list.insert(END,item)

def fillout(e):
    # borrar todo lo que haya en la entry
    my_entry.delete(0,END)
    # inserta el elemento seleccionado del listbox en el entry
    my_entry.insert(0,my_list.get(ACTIVE))

def check(e):
    # pide el valor de la entrada
    typed = my_entry.get()
    # si es un string vacio
    if typed == '':
        data = lreyes
        # la listbox permanece igual
    else:
        # se borra toda la lista de data
        data = []
        for item in lreyes:# para cada elemento de nuestra lista
            if typed.lower() in item.lower():
                # si la entrada en minusculas
                # es igual a un valor dentro de la lista en minusculas
                # se añade a la lista
                data.append(item)

    update(data)# MIUESTRA LOS VALORES ACTUALES

# Valores de los botones
def seleccionar():
    print(my_entry.get())

# Valores 
def nuevo_pedido(treeview_data):
    total_items 
    treeview.insert(parent=item[0], index=item[1], iid=item[2], text=item[3], values=item[4])

def selectbox(e):
    comboactual = menu_tiendas.get()
    if comboactual == 'Tienda de Reyes': # PRODUCTOS DEFINIDOS
     # PRODUCTOS DEFINIDOS 
        tienda1 = wb['Reyes']
        ltienda1 = loaddb(tienda1)
        update(ltienda1)
    elif comboactual == 'Carniceria':
        tienda2 = wb['Carniceria']
        ltienda2 = loaddb(tienda2)
        update(ltienda2)
    elif comboactual == 'Papeleria de José':
        tienda3 = wb['Papeleria de José']
        ltienda3 = loaddb(tienda3)
        update(ltienda3)
    else:
        pass
##    print(comboactual)
root = tk.Tk()
root.title("Forest")
root.option_add("*tearOff", False) # This is always a good idea

g = tk.DoubleVar(value=75.0)

# Make the app responsive
root.columnconfigure(index=0, weight=1)
root.columnconfigure(index=1, weight=1)
root.columnconfigure(index=2, weight=1)
root.rowconfigure(index=0, weight=1)
root.rowconfigure(index=1, weight=1)
root.rowconfigure(index=2, weight=1)

# Create a style
style = ttk.Style(root)

# Import the tcl file
root.tk.call("source", "forest-dark.tcl")

# Set the theme with the theme_use method
style.theme_use("forest-dark")

botones_frame = ttk.LabelFrame(root, text="Que desea hacer?", padding=(20, 10))
botones_frame.grid(row=0, column=0, padx=(20, 10), pady=(20, 10), sticky="nsew")

button = ttk.Button(botones_frame, text="Nueva orden",command=nuevo_pedido)
button.grid(row=0, column=0, padx=5, pady=10, sticky="nsew")

# Panedwindow
paned = ttk.PanedWindow(root)
paned.grid(row=1, column=1, pady=(25, 5), sticky="nsew", rowspan=3)

# Pane #1
pane_1 = ttk.Frame(paned)
paned.add(pane_1, weight=1)

# Create a Frame for the Treeview
treeFrame = ttk.Frame(pane_1)
treeFrame.pack(expand=True, fill="both", padx=5, pady=5)

# Scrollbar
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

# Treeview
treeview = ttk.Treeview(treeFrame, selectmode="extended", yscrollcommand=treeScroll.set, columns=(1, 2), height=12)
treeview.pack(expand=True, fill="both")
treeScroll.config(command=treeview.yview)

# Treeview columns
treeview.column("#0", width=120)
treeview.column(1, anchor="w", width=120)
treeview.column(2, anchor="w", width=120)

# Treeview headings
treeview.heading("#0", text="Column 1", anchor="center")
treeview.heading(1, text="Column 2", anchor="center")
treeview.heading(2, text="Column 3", anchor="center")

# Define treeview data
treeview_data = [
    ("", "end", 1, "Orden 1", ("", "")),
    (1, "end", 2, "Tienda reyes", ("", "")),
    (2, "end", 3, "producto 1", ("xxxx Kg", "$xx.xx")),
    (1, "end", 4, "Tienda Mandrake", ("", "")),
    (4, "end", 5, "producto 1", ("xxxx Kg", "$xx.xx")),
    ("", "end", 6, "Copias Edgar", ("", "")),
    (6, "end", 7, "Child", ("Subitem 2.1", "Value 2.1")),
    (6, "end", 8, "Sub-parent", ("Subitem 2.2", "Value 2.2")),
    (8, "end", 9, "Child", ("Subitem 2.2.1", "Value 2.2.1")),
    (8, "end", 10, "Child", ("Subitem 2.2.2", "Value 2.2.2")),
    (8, "end", 11, "Child", ("Subitem 2.2.3", "Value 2.2.3")),
    (6, "end", 12, "Child", ("Subitem 2.3", "Value 2.3")),
    (6, "end", 13, "Child", ("Subitem 2.4", "Value 2.4")),
    ]
    # Item
    # (parent, index, iid, text, values)

# Insert treeview data
for item in treeview_data:
    treeview.insert(parent=item[0], index=item[1], iid=item[2], text=item[3], values=item[4])
    if item[0] == "" or item[2] in (8, 12):
        treeview.item(item[2], open=True) # Open parents

# Select and scroll
treeview.selection_set(10)
treeview.see(7)

# Panedwindow
paned2 = ttk.PanedWindow(root)
paned2.grid(row=1, column=0, pady=(25, 5), sticky="nsew", rowspan=3)

# Pane #2
pane_2 = ttk.Frame(paned2)
paned2.add(pane_2, weight=3)

# Notebook
notebook = ttk.Notebook(pane_2)

# Tab #1
tab_1 = ttk.Frame(notebook)
tab_1.columnconfigure(index=0, weight=1)
tab_1.columnconfigure(index=1, weight=1)
tab_1.rowconfigure(index=0, weight=1)
tab_1.rowconfigure(index=1, weight=1)
notebook.add(tab_1, text="Productos")

# Fill_3


wb = load_workbook('Lista_Reyes.xlsx')
reyes = wb['Reyes']

def loaddb(reyes):
    m_row = reyes.max_row
##    print(reyes)
##    print(m_row)
    lreyes = []
    for i in range(1,m_row+1):
        cell_obj = reyes.cell(row=i, column=1)# Se obtiene el la celda por completp
        print(cell_obj.value)# se imprime el valor dentro de la celda
        lreyes.append(cell_obj.value)# se agrega el valor a la lista lreyes
    return lreyes

lreyes = loaddb(reyes)

# Fill_5

lista_tiendas = [
    'Tienda de Reyes', # PRODUCTOS DEFINIDOS
    'Carniceria', # PRODUCTOS DEFINIDOS 
    'Papeleria de José', # PRODUCTOS DEFINIDOS
    'Café internet Edgar Online', # SERVICIOS DE COPIAS
    'Antojitos CETIS 33', # PRODUCTOS DEFINIDOS
    'Liconsa',# PRODUCTO DEFINIDO
    'Tortilleria',# PRODUCTO DEFINIDO
    'Purificadora de Agua',# PRODUCTOS DEFINIDO
    'Plomeria Y Electricista Memo'# SERVICIOS
    'OXXO',# PRODUCTOS DEFINIDOS
    ]

Seleccionado = StringVar()
Seleccionado.set(lista_tiendas[0])


##my_label = Label(tab_1, text="Start Typing...", font=("Helvetica",14),fg="grey")
##my_label.grid(row=0, column=0, padx=(20, 10), pady=(20, 0), sticky="ew")

##sel_prod_frame = ttk.LabelFrame(tab_1, text="Checkbuttons", padding=(20, 10))
##sel_prod_frame.grid(row=1, column=0, padx=10, pady=(30, 10), sticky="nsew", rowspan=3)

# Combobox
menu_tiendas = ttk.Combobox(tab_1, values=lista_tiendas)
menu_tiendas.current(0)
menu_tiendas.grid(row=0, column=0, padx=5, pady=10,  sticky="ew")

my_entry = ttk.Entry(tab_1,font=("Helvetica",20))
my_entry.grid(row=1, column=0, padx=(10, 20), pady=(20, 0), sticky="ew")

my_list = Listbox(tab_1,width=50)
my_list.grid(row=2, column=0, padx=(10, 20), pady=(20, 0), sticky="ew")

button = ttk.Button(tab_1, text=">>",command=seleccionar)
button.grid(row=3, column=0, padx=5, pady=10, sticky="nsew")

# Button
##button = ttk.Button(tab_1, text="seleccionar Producto",command=seleccionar)
##button.grid(row=3, column=0, padx=5, pady=10, sticky="nsew")

##lreyes = ["Peperoni","Peppers","Mushrooms","Cheese","Onions","Ham","Taco"]


update(lreyes)

my_list.bind("<<ListboxSelect>>",fillout)

my_entry.bind("<KeyRelease>",check)

menu_tiendas.bind("<<ComboboxSelected>>",selectbox)

### Scale
##scale = ttk.Scale(tab_1, from_=100, to=0, variable=g, command=lambda event: g.set(scale.get()))
##scale.grid(row=0, column=0, padx=(20, 10), pady=(20, 0), sticky="ew")
##
### Progressbar
##progress = ttk.Progressbar(tab_1, value=0, variable=g, mode="determinate")
##progress.grid(row=0, column=1, padx=(10, 20), pady=(20, 0), sticky="ew")
##
### Label
##label = ttk.Label(tab_1, text="Forest ttk theme", justify="center")
##label.grid(row=1, column=0, pady=10, columnspan=2)

# Tab #2
tab_2 = ttk.Frame(notebook)
notebook.add(tab_2, text="Tab 2")

# Tab #3
tab_3 = ttk.Frame(notebook)
notebook.add(tab_3, text="Tab 3")

notebook.pack(expand=True, fill="both", padx=5, pady=5)

# Sizegrip
sizegrip = ttk.Sizegrip(root)
sizegrip.grid(row=100, column=100, padx=(0, 5), pady=(0, 5))

# Center the window, and set minsize
root.update()
root.minsize(root.winfo_width(), root.winfo_height())
x_cordinate = int((root.winfo_screenwidth()/2) - (root.winfo_width()/2))
y_cordinate = int((root.winfo_screenheight()/2) - (root.winfo_height()/2))
root.geometry("+{}+{}".format(x_cordinate, y_cordinate))

# Start the main loop
root.mainloop()
